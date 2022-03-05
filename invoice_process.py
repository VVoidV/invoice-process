# -*- coding: utf-8 -*-
import pdfplumber
import openpyxl
import os
import re
import shutil

ROOT = r"D:\Chinatelecom\test"
XML_PATH = "D:\\Chinatelecom\\test\\test_2.xlsx"
TRAVEL_PATH = r"D:\ChinaTelecom\test\Q\交通"
CELLPHONE_PATH = r"D:\ChinaTelecom\test\Q\通讯"
SHEET_NAME = "工作表1"

# 分配发票的单张金额
values = [1000, 900, 800, 700, 600, 500, 400, 300, 200, 100, 50]


class Need:
        def __init__(self, amount, step):
                self.amount = amount
                self.detail = self.split_amount(amount, step)

        def split_amount(self, amount, step):
                remain = amount
                result = {}
                if amount == 0:
                        return result
                for value in values:
                        if value < step:
                                print("illegal need")
                                raise

                        count = remain // value
                        if count > 0:
                                result[value] = count
                                remain -= count * value

                        if remain == 0:
                                break

                assert(remain == 0)
                return result

        def __repr__(self):
                return self.__str__()

        def __str__(self):
                res = " "
                for key, val in self.detail.items():
                        res += str(key) + " 元 " + str(val) + " 张,"
                return res


class Person:
        def __init__(self, name, travel, cellphone):
                self.name = name or ""
                self.travel = travel or 0
                self.cellphone = cellphone or 0
                self.travel_need = Need(self.travel, 100)

        def __repr__(self):
                return self.__str__()

        def __str__(self):
                return self.name + "\t交通:" + str(self.travel) + "\t通信: " + str(self.cellphone) + "\n\t交通发票: " + str(self.travel_need)


class Invoice:
        def __init__(self, path):
                self.path = path
                self.parse_pdf()

        # 将纵向位置相近的文字整理到一起
        def get_line_pack(self, words):
            lines = {}
            for word in words:
                top = int(word['top'])
                bottom = int(word['bottom'])
                pos = (top + bottom) // 2
                text = word['text']
                if pos not in lines:
                    lines[pos] = [text]
                else:
                    lines[pos].append(text)

            lines_pack = []
            last_pos = None
            for pos in sorted(lines):
                arr = lines[pos]

                if len(lines_pack) > 0 and pos - last_pos <= 2:
                    lines_pack[-1] += arr
                    continue

                lines_pack.append(arr)
                last_pos = pos
                continue
            return lines_pack

        def parse_pdf(self):
            if not self.path:
                return
            path = self.path
            # path = "D:\\ChinaTelecom\\test\\invoice_travel.pdf"

            with pdfplumber.open(path) as pdf:
                page = pdf.pages[0]
                pack = self.get_line_pack(page.extract_words())
                for p in pack:
                    for s in p:
                        if "发票号码" in s:
                            for q in p:
                                m = re.search("[0-9]{8}", q)
                                if m:
                                    id = m.group()
                                    break
                        elif "价税合计" in s:
                            for q in p:
                                m = re.search("(￥|¥)[0-9,\.]+|[0-9,\.]+", q)
                                if m:
                                    num = m.group().strip('￥').strip("¥")
                                    amount = float(num)
                                    break

                self.id = id
                self.amount = amount


        def __repr__(self):
                return self.__str__()

        def __str__(self):
                return str(self.id) + ", 金额:" + str(self.amount) + " 路径: " + str(self.path)


class Dispather:
        def __init__(self, root, xlsx_path, travel_path, cellphone_path):
                self.root = root
                self.xml_path = xlsx_path
                self.travel_path = travel_path
                self.cellphone_path = cellphone_path
                self.employees = {}
                self.travel_invoices = []
                self.cellphone_invoices = []
                self.sheet_name = SHEET_NAME

        def load_xlsx(self):
                wb = openpyxl.load_workbook(self.xml_path, data_only=True)
                sheet = wb["工作表1"]

                # find colume
                for col in sheet.iter_cols():
                        if col[0].value == "姓名":
                                name_idx = col[0].col_idx
                        elif col[0].value == "交通":
                                travel_idx = col[0].col_idx
                        elif col[0].value == "通信":
                                phone_idx = col[0].col_idx
                        else:
                                continue

                for row in sheet.iter_rows(min_row=2):
                        name = row[name_idx - 1].value
                        if not name:
                                continue
                        travel = row[travel_idx - 1].value
                        phone = row[phone_idx - 1].value
                        p = Person(name, travel, phone)
                        self.employees[name] = p
                        print(p)
                        print('='*60)

        def load_invoices(self, path):
            pdfs = []
            res = []
            fail_path = []
            for root, _, files in os.walk(path):
                for fn in files:
                    ext = os.path.splitext(fn)[1].lower()
                    if ext != '.pdf':
                        continue
                    fpth = os.path.join(root, fn)
                   # print(f'pdf: {fpth}')
                    pdfs.append(fpth)

            for pdf_path in pdfs:
                # print(f"processing {pdf_path}")
                try:
                    invoice = Invoice(pdf_path)
                except:
                    print(f'{pdf_path} failed to parse')
                    fail_path.append(pdf_path)
                    continue

                res.append(invoice)
            
            return (res, fail_path)
        
        def load_all_invoices(self):
            if self.cellphone_path:
                self.cellphone_invoices, self.fail_cellphone = self.load_invoices(self.cellphone_path)
            if self.cellphone_path:
                self.travel_invoices, self.fail_travel = self.load_invoices(self.travel_path)
        
        def sort_invoices(self):
                # self.cellphone_sorted = dict(zip(values, [[]]*len(values)))
                # self.travel_sorted = dict(zip(values, [[]]*len(values)))
                self.cellphone_sorted = {}
                self.travel_sorted = {}
                for ino in self.cellphone_invoices:
                        if ino.amount not in self.cellphone_sorted:
                                self.cellphone_sorted[ino.amount] = []
                        self.cellphone_sorted[ino.amount].append(ino)

                for ino in self.travel_invoices:
                        if ino.amount not in self.travel_sorted:
                                self.travel_sorted[ino.amount] = []
                        self.travel_sorted[ino.amount].append(ino)
                
                for key, val in self.cellphone_sorted.items():
                        print(f'{key}, count: {len(val)}')

                print("travel", "="*60)
                for key, val in self.travel_sorted.items():
                        print(f'{key}, count: {len(val)}')



        def get_one(self, val, pool):
                try:
                        arr = pool[val]
                        invoice = arr.pop(0)
                        return invoice
                except:
                        return None
        
        def gen_name(self, name, type, idx, amount):
                return name + "-" + type + "-" + str(idx) + "-" + str(amount)


        def dispatch(self):
                travel_res_path = os.path.join(self.root, "交通分配")
                phone_res_path = os.path.join(self.root, "通信分配")

                for emp in self.employees.values():
                        dest = os.path.join(travel_res_path, emp.name)
                        emp.job = []
                        emp.travel_miss = []
                        emp.phone_miss = []
                        detail = emp.travel_need.detail
                        #分配交通票
                        index = 0
                        for val,count in detail.items():
                                while count > 0:
                                        inv = self.get_one(val, self.travel_sorted)

                                        if inv == None:
                                                print(f'{emp.name} get 交通 {val} failed')
                                                emp.travel_miss.append(val)
                                                count -=1
                                                continue
                                                
                                        name = self.gen_name(emp.name, "交通", index, inv.amount)
                                        os.path.join(dest, name)
                                        print(f'{emp.name} got {inv.path}')
                                        emp.job.append((inv.path, dest))
                                        count-=1
                                        index += 1
                        # 分配通信票
                        phone_val = emp.cellphone
                        inv = self.get_one(phone_val, self.cellphone_sorted)
                        if inv == None:
                                print(f'{emp.name} get 通信 {val} failed')
                                emp.phone_miss.append(val)
                                continue

                        name = self.gen_name(emp.name, "通信", index, inv.amount)
                        dest = os.path.join(phone_res_path, emp.name)
                        dest = os.path.join(dest, name)
                        emp.job.append((inv.path, dest))



        def count_need(self):
                travel_need = dict(zip(values, [0]*len(values)))
                req_travel_total = 0
                for entry in self.employees.values():
                        req_travel_total += entry.travel
                        for key, value in entry.travel_need.detail.items():
                                travel_need[key] += value
                
                travel_total = 0
                for key, val in travel_need.items():
                        travel_total += key * val

                print("分配后交通总额: ", travel_total)
                print("申请交通总额: ", req_travel_total)
                assert(travel_total == req_travel_total)
                print(travel_need)
                






        


        
if __name__ == "__main__":
        worker = Dispather(ROOT, XML_PATH, TRAVEL_PATH, CELLPHONE_PATH)                
        worker.load_xlsx()
        worker.count_need()
        worker.load_all_invoices()
        worker.sort_invoices()

        worker.dispatch()
        pass

